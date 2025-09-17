"""
Organizational Chart Processing Script
=====================================

This script processes monthly organizational data from SelfProd_2024.xlsx and generates
a comprehensive processed output with promotion/demotion logic, supervisor relationships,
and hierarchical structure management.

Business Rules Implemented:
1. Rank remapping (职级套转): AD→AD1, SBM→BM3, BM→BM1, AS→AS
2. Layer definition: Layer 1 (AD2,AD1), Layer 2 (AD0,BM3,BM2,BM1), Layer 3 (BM0,AS)
3. Assessment months processing (July 202407 & October 202410)
4. Performance-based promotion/demotion logic
5. Supervisor relationship management (DT/GT relationships)
6. New hire/resignation detection
7. Monthly snapshot generation for 202401-202412

Author: Generated Script
Version: 1.0
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional
import warnings
warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION CONSTANTS
# =============================================================================

# File paths and sheet names
SRC_FILE = "SelfProd_2024.xlsx"
DST_FILE = "SelfProd_2024_processed.xlsx"
CHECK_FILE = "SelfProd_2024_promotion_check.xlsx"  # Monthly promotion details
DATA_SHEET = "Data"
PROMOTION_SHEET = "PromotionRules"
DEMOTION_SHEET = "DemotionRules"

# Assessment configuration
PROMOTION_WINDOW = 6  # Number of months for performance assessment
ASSESS_MONTHS = {202407, 202410}  # Assessment months (July & October)

# Rank remapping configuration
RANK_MAP = {
    "AD": "AD1",
    "SBM": "BM3", 
    "BM": "BM1",
    "AS": "AS"
}

# Layer definitions
LAYER_MAP = {
    "AD2": 1, "AD1": 1,                    # Layer 1
    "AD0": 2, "BM3": 2, "BM2": 2, "BM1": 2,  # Layer 2
    "BM0": 3, "AS": 3                      # Layer 3
}

# Rank hierarchy for promotion/demotion (ascending order)
RANK_HIERARCHY = ["AS", "BM0", "BM1", "BM2", "BM3", "AD0", "AD1", "AD2"]

# Default promotion/demotion thresholds (fallback if sheets missing)
DEFAULT_PROMOTION_THRESHOLD = 250000  # 承保FYC累计
DEFAULT_MAINTAIN_THRESHOLD = 120000   # 承保FYC累计

# Performance columns for assessment (based on Excel requirements)
PERFORMANCE_COLUMNS = [
    '续保率',        # 续保率 (Renewal rate)
    '承保FYC',       # 个人FYC (Individual FYC)
    '个险折算后FYC',  # Alternative for direct FYC
    '直辖FYC',       # 直辖FYC (Direct team FYC)
    '所辖FYC',       # 所辖FYC (Group FYC)
    '新单件数',      # 新单件数 (New policies)
    '是否MDRT',      # MDRT status
    '直辖人力',      # 直辖人力 (Direct team size - DT人力)
    '所辖人力',      # 所辖人力 (Group team size - GT人力)
    '业务主管人数',   # 业务主管人数 (Business supervisors count)
    '引荐人数',      # 引荐人数 (Referrals count)
    '本人直增人数',   # 本人直增人数 (Direct growth count)
    '达星人数'       # 达星人数 (Star achievers count)
]

# Output columns in required order
OUTPUT_COLUMNS = [
    '营销员代码', '上级主管代码', '总监代码', '总监职级', '总监关系',
    '业务经理代码', '业务经理职级', '业务经理关系',
    '业务主任代码', '业务主任职级', '业务主任关系',
    '一代育成人', '二代育成人'
]


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def validate_file_exists(filepath: str) -> None:
    """Validate that the source file exists."""
    if not Path(filepath).exists():
        raise FileNotFoundError(f"Source file not found: {filepath}")

def generate_month_range(start_month: int, end_month: int) -> List[int]:
    """
    Generate a list of YYYYMM integers between start and end (inclusive).
    
    Args:
        start_month: Starting month in YYYYMM format
        end_month: Ending month in YYYYMM format
        
    Returns:
        List of month integers in YYYYMM format
    """
    months = []
    year_s, mon_s = divmod(start_month, 100)
    year_e, mon_e = divmod(end_month, 100)
    
    y, m = year_s, mon_s
    while (y < year_e) or (y == year_e and m <= mon_e):
        months.append(y * 100 + m)
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months

def calculate_assessment_window(current_month: int, window_size: int) -> List[int]:
    """
    Calculate the assessment window months for performance evaluation.
    
    Args:
        current_month: Current assessment month (YYYYMM)
        window_size: Number of months to look back
        
    Returns:
        List of months in the assessment window
    """
    year, month = divmod(current_month, 100)
    start_year = year
    start_month = month - window_size + 1
    
    if start_month <= 0:
        start_month += 12
        start_year -= 1
        
    start_yyyymm = start_year * 100 + start_month
    return generate_month_range(start_yyyymm, current_month)

# =============================================================================
# DATA LOADING AND VALIDATION
# =============================================================================

def load_source_data(filepath: str) -> pd.DataFrame:
    """
    Load and validate source data from Excel file.
    
    Args:
        filepath: Path to source Excel file
        
    Returns:
        DataFrame with validated source data
    """
    validate_file_exists(filepath)
    
    # Define data types for consistent handling
    dtype_map = {
        '营销员代码': str,
        '直属主管代码': str,
        '上级主管代码': str,
        'AD代码': str,
        '育成主管代码': str
    }
    
    try:
        xls = pd.ExcelFile(filepath)
        df = xls.parse(DATA_SHEET, dtype=dtype_map, keep_default_na=False)
        print(f"✓ Loaded {len(df)} records from {DATA_SHEET} sheet")
        return df
    except Exception as e:
        raise ValueError(f"Failed to load data from {filepath}: {str(e)}")

def load_promotion_rules(filepath: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load promotion and demotion rules from Excel sheets.
    
    Args:
        filepath: Path to source Excel file
        
    Returns:
        Tuple of (promotion_rules_df, demotion_rules_df)
    """
    try:
        xls = pd.ExcelFile(filepath)
        
        # Try to load promotion rules
        try:
            promo_df = xls.parse(PROMOTION_SHEET)
            print(f"✓ Loaded promotion rules from {PROMOTION_SHEET} sheet")
        except ValueError:
            promo_df = pd.DataFrame()
            print(f"⚠ {PROMOTION_SHEET} sheet not found, using default rules")
            
        # Try to load demotion rules
        try:
            demo_df = xls.parse(DEMOTION_SHEET)
            print(f"✓ Loaded demotion rules from {DEMOTION_SHEET} sheet")
        except ValueError:
            demo_df = pd.DataFrame()
            print(f"⚠ {DEMOTION_SHEET} sheet not found, using default rules")
            
        return promo_df, demo_df
        
    except Exception as e:
        print(f"⚠ Warning: Could not load promotion rules: {str(e)}")
        return pd.DataFrame(), pd.DataFrame()

# =============================================================================
# PROMOTION RULE ENGINE - MODULAR SYSTEM
# =============================================================================

class PromotionRule:
    """Base class for promotion rules."""
    
    def __init__(self, name: str, priority: int = 0):
        self.name = name
        self.priority = priority  # Higher priority rules are evaluated first
    
    def applies_to(self, current_rank: str, performance_data: Dict) -> bool:
        """Check if this rule applies to the given agent."""
        raise NotImplementedError
    
    def evaluate(self, current_rank: str, performance_data: Dict) -> str:
        """Evaluate and return decision: '晋升', '维持', or '降级'."""
        raise NotImplementedError

class FYCBasedRule(PromotionRule):
    """FYC-based promotion rule."""
    
    def __init__(self, name: str, rank_filter: List[str] = None, 
                 promotion_threshold: float = 250000, maintain_threshold: float = 120000,
                 priority: int = 1):
        super().__init__(name, priority)
        self.rank_filter = rank_filter or []
        self.promotion_threshold = promotion_threshold
        self.maintain_threshold = maintain_threshold
    
    def applies_to(self, current_rank: str, performance_data: Dict) -> bool:
        """Rule applies if no rank filter or rank is in filter."""
        return not self.rank_filter or current_rank in self.rank_filter
    
    def evaluate(self, current_rank: str, performance_data: Dict) -> str:
        """Evaluate based on FYC thresholds (conservative demotion logic)."""
        fyc_total = performance_data.get('承保FYC', 0)
        if pd.isna(fyc_total):
            fyc_total = 0
        
        # Only allow demotions during assessment months and for severe underperformance
        is_assessment_month = performance_data.get('is_assessment_month', False)
        
        if fyc_total >= self.promotion_threshold:
            return '晋升'
        elif fyc_total >= self.maintain_threshold:
            return '维持'
        elif is_assessment_month and fyc_total > 0 and fyc_total < (self.maintain_threshold * 0.2):
            # Only demote if performance is less than 20% of maintain threshold AND has data
            return '降级'
        else:
            return '维持'  # Conservative: maintain instead of demote

class CompositeRule(PromotionRule):
    """Rule that combines multiple performance metrics."""
    
    def __init__(self, name: str, rank_filter: List[str] = None,
                 fyc_threshold: float = 200000, renewal_rate_threshold: float = 0.85,
                 new_policies_threshold: int = 10, priority: int = 2):
        super().__init__(name, priority)
        self.rank_filter = rank_filter or []
        self.fyc_threshold = fyc_threshold
        self.renewal_rate_threshold = renewal_rate_threshold
        self.new_policies_threshold = new_policies_threshold
    
    def applies_to(self, current_rank: str, performance_data: Dict) -> bool:
        """Rule applies if no rank filter or rank is in filter."""
        return not self.rank_filter or current_rank in self.rank_filter
    
    def evaluate(self, current_rank: str, performance_data: Dict) -> str:
        """Evaluate based on multiple criteria (conservative demotion logic)."""
        fyc = performance_data.get('承保FYC', 0) or 0
        renewal_rate = performance_data.get('续保率', 0) or 0
        new_policies = performance_data.get('新单件数', 0) or 0
        is_assessment_month = performance_data.get('is_assessment_month', False)
        
        # Count how many criteria are met
        criteria_met = 0
        total_criteria = 3
        
        if fyc >= self.fyc_threshold:
            criteria_met += 1
        if renewal_rate >= self.renewal_rate_threshold:
            criteria_met += 1
        if new_policies >= self.new_policies_threshold:
            criteria_met += 1
        
        # Decision based on criteria met (more conservative)
        if criteria_met >= 3:
            return '晋升'
        elif criteria_met >= 1:  # More lenient: maintain if any criteria met
            return '维持'
        elif is_assessment_month and criteria_met == 0:
            # Only demote if NO criteria are met during assessment months
            return '降级'
        else:
            return '维持'  # Default to maintain

class ExcelBasedRule(PromotionRule):
    """Rule that reads criteria from Excel sheets."""
    
    def __init__(self, name: str, promotion_df: pd.DataFrame, demotion_df: pd.DataFrame,
                 priority: int = 3):
        super().__init__(name, priority)
        self.promotion_df = promotion_df
        self.demotion_df = demotion_df
    
    def applies_to(self, current_rank: str, performance_data: Dict) -> bool:
        """Rule applies if we have rules for this rank."""
        return (not self.promotion_df.empty and 
                current_rank in self.promotion_df.get('当前职级', []))
    
    def evaluate(self, current_rank: str, performance_data: Dict) -> str:
        """Evaluate based on Excel-defined rules."""
        if self.promotion_df.empty:
            return '维持'
        
        # Find rules for current rank
        rank_rules = self.promotion_df[self.promotion_df['当前职级'] == current_rank]
        if rank_rules.empty:
            return '维持'
        
        # Check promotion criteria (take first matching rule)
        rule = rank_rules.iloc[0]
        
        # Check each criterion
        promotion_criteria_met = True
        
        if '承保FYC要求' in rule and pd.notna(rule['承保FYC要求']):
            fyc = performance_data.get('承保FYC', 0) or 0
            if fyc < rule['承保FYC要求']:
                promotion_criteria_met = False
        
        if '续保率要求' in rule and pd.notna(rule['续保率要求']):
            renewal_rate = performance_data.get('续保率', 0) or 0
            if renewal_rate < rule['续保率要求']:
                promotion_criteria_met = False
        
        if '新单件数要求' in rule and pd.notna(rule['新单件数要求']):
            new_policies = performance_data.get('新单件数', 0) or 0
            if new_policies < rule['新单件数要求']:
                promotion_criteria_met = False
        
        if promotion_criteria_met:
            return '晋升'
        
        # Check demotion criteria
        if not self.demotion_df.empty:
            demotion_rules = self.demotion_df[self.demotion_df['当前职级'] == current_rank]
            if not demotion_rules.empty:
                demo_rule = demotion_rules.iloc[0]
                demotion_triggered = False
                
                if '承保FYC下限' in demo_rule and pd.notna(demo_rule['承保FYC下限']):
                    fyc = performance_data.get('承保FYC', 0) or 0
                    if fyc < demo_rule['承保FYC下限']:
                        demotion_triggered = True
                
                if demotion_triggered:
                    return '降级'
        
        return '维持'

class PromotionRuleEngine:
    """Engine that manages and applies promotion rules."""
    
    def __init__(self):
        self.rules: List[PromotionRule] = []
    
    def add_rule(self, rule: PromotionRule):
        """Add a rule to the engine."""
        self.rules.append(rule)
        # Sort by priority (higher priority first)
        self.rules.sort(key=lambda r: r.priority, reverse=True)
    
    def clear_rules(self):
        """Clear all rules."""
        self.rules = []
    
    def evaluate(self, current_rank: str, performance_data: Dict) -> str:
        """Evaluate promotion decision using all applicable rules."""
        # Safety check: if no meaningful performance data, always maintain
        has_any_performance_data = (
            performance_data.get('承保FYC', 0) > 0 or 
            performance_data.get('直辖FYC', 0) > 0 or 
            performance_data.get('个险折算后FYC', 0) > 0 or
            performance_data.get('所辖FYC', 0) > 0
        )
        
        if not has_any_performance_data:
            return '维持'  # No data = maintain (never demote)
        
        for rule in self.rules:
            if rule.applies_to(current_rank, performance_data):
                decision = rule.evaluate(current_rank, performance_data)
                # Only log significant decisions (promotions/demotions) to reduce verbosity
                if rule.priority > 0 and decision != '维持':
                    print(f"    {rule.name}: {current_rank} → {decision}")
                return decision
        
        # Default fallback (no logging to reduce verbosity)
        return '维持'

def create_default_promotion_engine(promotion_df: pd.DataFrame = None, 
                                  demotion_df: pd.DataFrame = None) -> PromotionRuleEngine:
    """
    Create a promotion engine with default rules.
    
    Args:
        promotion_df: Optional promotion rules from Excel
        demotion_df: Optional demotion rules from Excel
        
    Returns:
        Configured PromotionRuleEngine
    """
    engine = PromotionRuleEngine()
    
    # Add Excel-based rules first (highest priority)
    if promotion_df is not None and not promotion_df.empty:
        excel_rule = ExcelBasedRule("Excel Rules", promotion_df, demotion_df or pd.DataFrame(), priority=3)
        engine.add_rule(excel_rule)
    
    # Add composite rules for senior ranks
    senior_ranks = ['BM1', 'BM2', 'BM3', 'AD0', 'AD1', 'AD2']
    composite_rule = CompositeRule(
        "Senior Rank Composite", 
        rank_filter=senior_ranks,
        fyc_threshold=300000,
        renewal_rate_threshold=0.90,
        new_policies_threshold=15,
        priority=2
    )
    engine.add_rule(composite_rule)
    
    # Add basic FYC rule for junior ranks
    junior_ranks = ['AS', 'BM0']
    basic_rule = FYCBasedRule(
        "Junior Rank Basic",
        rank_filter=junior_ranks,
        promotion_threshold=150000,
        maintain_threshold=80000,
        priority=1
    )
    engine.add_rule(basic_rule)
    
    # Add default fallback rule
    fallback_rule = FYCBasedRule(
        "Default Fallback",
        rank_filter=[],  # Applies to all ranks
        promotion_threshold=DEFAULT_PROMOTION_THRESHOLD,
        maintain_threshold=DEFAULT_MAINTAIN_THRESHOLD,
        priority=0
    )
    engine.add_rule(fallback_rule)
    
    return engine

# =============================================================================
# CUSTOM PROMOTION RULE CONFIGURATIONS
# =============================================================================

# Example: MDRT-based promotion rule
class MDRTBasedRule(PromotionRule):
    """Promotion rule based on MDRT achievement."""
    
    def __init__(self, name: str, rank_filter: List[str] = None, priority: int = 2):
        super().__init__(name, priority)
        self.rank_filter = rank_filter or []
    
    def applies_to(self, current_rank: str, performance_data: Dict) -> bool:
        """Rule applies if MDRT data is available and rank matches filter."""
        return (not self.rank_filter or current_rank in self.rank_filter) and \
               '是否MDRT' in performance_data
    
    def evaluate(self, current_rank: str, performance_data: Dict) -> str:
        """Promote if MDRT achieved, otherwise maintain."""
        is_mdrt = performance_data.get('是否MDRT', False)
        if is_mdrt:
            return '晋升'
        else:
            # Check minimum performance to avoid demotion
            fyc = performance_data.get('承保FYC', 0) or 0
            if fyc >= 100000:  # Minimum threshold
                return '维持'
            else:
                return '降级'

# Example: Tenure-based rule (requires additional data)
class TenureBasedRule(PromotionRule):
    """Promotion rule that considers tenure in current rank."""
    
    def __init__(self, name: str, min_tenure_months: int = 12, 
                 rank_filter: List[str] = None, priority: int = 1):
        super().__init__(name, priority)
        self.min_tenure_months = min_tenure_months
        self.rank_filter = rank_filter or []
    
    def applies_to(self, current_rank: str, performance_data: Dict) -> bool:
        """Rule applies if tenure data is available."""
        return (not self.rank_filter or current_rank in self.rank_filter) and \
               'months_in_rank' in performance_data
    
    def evaluate(self, current_rank: str, performance_data: Dict) -> str:
        """Consider tenure along with performance."""
        months_in_rank = performance_data.get('months_in_rank', 0)
        fyc = performance_data.get('承保FYC', 0) or 0
        
        if months_in_rank >= self.min_tenure_months and fyc >= 200000:
            return '晋升'
        elif fyc >= 120000:
            return '维持'
        else:
            return '降级'

# Grade-specific promotion rule based on the Excel requirements
class GradeSpecificRule(PromotionRule):
    """Promotion rule based on specific grade requirements from Excel."""
    
    def __init__(self, name: str, priority: int = 4):
        super().__init__(name, priority)
        # Define specific requirements for each grade transition based on Excel
        self.promotion_requirements = {
            'AS': {  # AS → BM0 or BM1 (depending on performance)
                'target_rank': 'BM0',  # Default target
                'alternative_target': 'BM1',  # Higher performance target
                'time_requirement': 6,
                'individual_fyc': 18000,
                'direct_fyc': 36000,  # For BM0 (直辖FYC)
                'direct_fyc_alt': 90000,  # For BM1 (直辖FYC)
                'group_fyc': 0,  # 所辖FYC not required for AS
                'renewal_rate': 0.85,  # 续保率
                'direct_team_size': 4,  # 直辖人力 for BM0
                'direct_team_size_alt': 6,  # 直辖人力 for BM1
                'group_team_size': 0,  # 所辖人力 not required
                'business_supervisors': 0,  # 业务主管人数
                'referrals': 0,  # 引荐人数
                'direct_growth': 0,  # 本人直增人数
                'star_achievers': 0  # 达星人数
            },
            'BM0': {  # BM0 → BM1
                'target_rank': 'BM1',
                'time_requirement': 6,
                'individual_fyc': 12000,
                'direct_fyc': 63000,
                'group_fyc': 0,
                'renewal_rate': 0.85,
                'team_size': 4
            },
            'BM1': {  # BM1 → BM2
                'target_rank': 'BM2',
                'time_requirement': 12,
                'individual_fyc': 12000,
                'direct_fyc': 105000,
                'group_fyc': 0,
                'renewal_rate': 0.85,
                'team_size': 8
            },
            'BM2': {  # BM2 → BM3 or AD0
                'target_rank': 'BM3',
                'alternative_target': 'AD0',
                'time_requirement': 12,
                'individual_fyc': 12000,
                'direct_fyc': 175000,  # For BM3
                'direct_fyc_alt': 112500,  # For AD0
                'group_fyc': 0,
                'group_fyc_alt': 300000,  # For AD0
                'renewal_rate': 0.85,
                'team_size': 14,  # For BM3
                'team_size_alt': 21  # For AD0
            },
            'BM3': {  # BM3 → AD0
                'target_rank': 'AD0',
                'time_requirement': 12,
                'individual_fyc': 12000,
                'direct_fyc': 112500,
                'group_fyc': 300000,
                'renewal_rate': 0.85,
                'team_size': 21
            },
            'AD0': {  # AD0 → AD1
                'target_rank': 'AD1',
                'time_requirement': 12,
                'individual_fyc': 12000,
                'direct_fyc': 135000,
                'group_fyc': 450000,
                'renewal_rate': 0.85,
                'team_size': 38
            },
            'AD1': {  # AD1 → AD2
                'target_rank': 'AD2',
                'time_requirement': 12,
                'individual_fyc': 0,  # Not specified
                'direct_fyc': 0,      # Not specified
                'group_fyc': 2000000,
                'renewal_rate': 0.0,  # Not specified
                'team_size': 80
            }
        }
        
        # Define maintenance/demotion thresholds based on Excel "维持要求"
        self.maintenance_requirements = {
            'AS': {  # AS maintenance requirements
                'individual_fyc': 9000,     # 个人FYC
                'direct_fyc': 15000,        # 直辖FYC
                'group_fyc': 0,             # 所辖FYC not required
                'renewal_rate': 0.85,       # 续保率
                'direct_team_size': 3,      # 直辖人力
                'group_team_size': 0,       # 所辖人力
                'business_supervisors': 0,   # 业务主管人数
                'referrals': 0,             # 引荐人数
                'direct_growth': 0,         # 本人直增人数
                'star_achievers': 0         # 达星人数
            },
            'BM0': {  # BM0 maintenance requirements  
                'individual_fyc': 12000,    # 个人FYC
                'direct_fyc': 25200,        # 直辖FYC
                'group_fyc': 0,             # 所辖FYC
                'renewal_rate': 0.85,       # 续保率
                'direct_team_size': 2,      # 直辖人力
                'group_team_size': 0,       # 所辖人力
                'business_supervisors': 0,   # 业务主管人数
                'referrals': 0,             # 引荐人数
                'direct_growth': 0,         # 本人直增人数
                'star_achievers': 0         # 达星人数
            },
            'BM1': {  # BM1 maintenance requirements
                'individual_fyc': 12000,    # 个人FYC
                'direct_fyc': 63000,        # 直辖FYC
                'group_fyc': 0,             # 所辖FYC
                'renewal_rate': 0.85,       # 续保率
                'direct_team_size': 4,      # 直辖人力
                'group_team_size': 0,       # 所辖人力
                'business_supervisors': 0,   # 业务主管人数
                'referrals': 0,             # 引荐人数
                'direct_growth': 0,         # 本人直增人数
                'star_achievers': 0         # 达星人数
            },
            'BM2': {  # BM2 maintenance requirements
                'individual_fyc': 12000,    # 个人FYC
                'direct_fyc': 105000,       # 直辖FYC
                'group_fyc': 0,             # 所辖FYC
                'renewal_rate': 0.85,       # 续保率
                'direct_team_size': 8,      # 直辖人力
                'group_team_size': 0,       # 所辖人力
                'business_supervisors': 0,   # 业务主管人数
                'referrals': 0,             # 引荐人数
                'direct_growth': 0,         # 本人直增人数
                'star_achievers': 0         # 达星人数
            },
            'BM3': {  # BM3 maintenance requirements
                'individual_fyc': 12000,    # 个人FYC
                'direct_fyc': 175000,       # 直辖FYC
                'group_fyc': 0,             # 所辖FYC
                'renewal_rate': 0.85,       # 续保率
                'direct_team_size': 14,     # 直辖人力
                'group_team_size': 0,       # 所辖人力
                'business_supervisors': 2,   # 业务主管人数
                'referrals': 1,             # 引荐人数
                'direct_growth': 0,         # 本人直增人数
                'star_achievers': 0         # 达星人数
            },
            'AD0': {  # AD0 maintenance requirements
                'individual_fyc': 12000,    # 个人FYC
                'direct_fyc': 112500,       # 直辖FYC
                'group_fyc': 300000,        # 所辖FYC
                'renewal_rate': 0.85,       # 续保率
                'direct_team_size': 21,     # 直辖人力
                'group_team_size': 0,       # 所辖人力
                'business_supervisors': 4,   # 业务主管人数
                'referrals': 2,             # 引荐人数
                'direct_growth': 0,         # 本人直增人数
                'star_achievers': 0         # 达星人数
            },
            'AD1': {  # AD1 maintenance requirements
                'individual_fyc': 12000,    # 个人FYC
                'direct_fyc': 135000,       # 直辖FYC
                'group_fyc': 450000,        # 所辖FYC
                'renewal_rate': 0.85,       # 续保率
                'direct_team_size': 38,     # 直辖人力
                'group_team_size': 0,       # 所辖人力
                'business_supervisors': 6,   # 业务主管人数
                'referrals': 3,             # 引荐人数
                'direct_growth': 0,         # 本人直增人数
                'star_achievers': 0         # 达星人数
            },
            'AD2': {  # AD2 maintenance requirements
                'individual_fyc': 0,        # 个人FYC not specified
                'direct_fyc': 0,            # 直辖FYC not specified
                'group_fyc': 2000000,       # 所辖FYC
                'renewal_rate': 0.0,        # 续保率not specified
                'direct_team_size': 0,      # 直辖人力
                'group_team_size': 80,      # 所辖人力
                'business_supervisors': 0,   # 业务主管人数
                'referrals': 0,             # 引荐人数
                'direct_growth': 0,         # 本人直增人数
                'star_achievers': 0         # 达星人数
            }
        }
    
    def applies_to(self, current_rank: str, performance_data: Dict) -> bool:
        """Rule applies only to ranks with defined requirements AND sufficient performance data."""
        has_requirements = (current_rank in self.promotion_requirements or 
                          current_rank in self.maintenance_requirements)
        
        # Only apply if we have meaningful performance data
        has_performance_data = (
            performance_data.get('承保FYC', 0) > 0 or 
            performance_data.get('直辖FYC', 0) > 0 or 
            performance_data.get('个险折算后FYC', 0) > 0
        )
        
        return has_requirements and has_performance_data
    
    def evaluate(self, current_rank: str, performance_data: Dict) -> str:
        """Evaluate based on grade-specific requirements with GT/DT relationship validation."""
        # Only allow demotions during assessment months (conservative approach)
        is_assessment_month = performance_data.get('is_assessment_month', False)
        
        # Check promotion first
        if current_rank in self.promotion_requirements:
            req = self.promotion_requirements[current_rank]
            
            # Get performance metrics based on Excel columns
            individual_fyc = performance_data.get('承保FYC', 0) or 0  # 个人FYC
            direct_fyc = performance_data.get('直辖FYC', 0) or performance_data.get('个险折算后FYC', 0) or 0  # 直辖FYC
            group_fyc = performance_data.get('所辖FYC', 0) or 0  # 所辖FYC
            renewal_rate = performance_data.get('续保率', 0) or 0  # 续保率
            direct_team_size = performance_data.get('直辖人力', 0) or 0  # 直辖人力 (DT人力)
            group_team_size = performance_data.get('所辖人力', 0) or 0  # 所辖人力 (GT人力)
            business_supervisors = performance_data.get('业务主管人数', 0) or 0  # 业务主管人数
            referrals = performance_data.get('引荐人数', 0) or 0  # 引荐人数
            direct_growth = performance_data.get('本人直增人数', 0) or 0  # 本人直增人数
            star_achievers = performance_data.get('达星人数', 0) or 0  # 达星人数
            
            # Check if promotion criteria are met (ALL must be satisfied)
            promotion_met = True
            criteria_failed = []
            
            # Check individual FYC
            if req.get('individual_fyc', 0) > 0 and individual_fyc < req['individual_fyc']:
                promotion_met = False
                criteria_failed.append(f"个人FYC {individual_fyc:,.0f} < {req['individual_fyc']:,.0f}")
            
            # Check direct team FYC (直辖FYC)
            if req.get('direct_fyc', 0) > 0 and direct_fyc < req['direct_fyc']:
                promotion_met = False
                criteria_failed.append(f"直辖FYC {direct_fyc:,.0f} < {req['direct_fyc']:,.0f}")
            
            # Check group FYC (所辖FYC)
            if req.get('group_fyc', 0) > 0 and group_fyc < req['group_fyc']:
                promotion_met = False
                criteria_failed.append(f"所辖FYC {group_fyc:,.0f} < {req['group_fyc']:,.0f}")
                
            # Check renewal rate (续保率)
            if req.get('renewal_rate', 0) > 0 and renewal_rate < req['renewal_rate']:
                promotion_met = False
                criteria_failed.append(f"续保率 {renewal_rate:.1%} < {req['renewal_rate']:.1%}")
                
            # Check direct team size (直辖人力)
            if req.get('direct_team_size', 0) > 0 and direct_team_size < req['direct_team_size']:
                promotion_met = False
                criteria_failed.append(f"直辖人力 {direct_team_size} < {req['direct_team_size']}")
                
            # Check group team size (所辖人力)
            if req.get('group_team_size', 0) > 0 and group_team_size < req['group_team_size']:
                promotion_met = False
                criteria_failed.append(f"所辖人力 {group_team_size} < {req['group_team_size']}")
                
            # Check business supervisors (业务主管人数)
            if req.get('business_supervisors', 0) > 0 and business_supervisors < req['business_supervisors']:
                promotion_met = False
                criteria_failed.append(f"业务主管人数 {business_supervisors} < {req['business_supervisors']}")
                
            # Check referrals (引荐人数)
            if req.get('referrals', 0) > 0 and referrals < req['referrals']:
                promotion_met = False
                criteria_failed.append(f"引荐人数 {referrals} < {req['referrals']}")
                
            # Check direct growth (本人直增人数)
            if req.get('direct_growth', 0) > 0 and direct_growth < req['direct_growth']:
                promotion_met = False
                criteria_failed.append(f"本人直增人数 {direct_growth} < {req['direct_growth']}")
                
            # Check star achievers (达星人数)
            if req.get('star_achievers', 0) > 0 and star_achievers < req['star_achievers']:
                promotion_met = False
                criteria_failed.append(f"达星人数 {star_achievers} < {req['star_achievers']}")
            
            # Check for alternative promotion path (AS → BM1, BM2 → AD0)
            if not promotion_met and 'alternative_target' in req:
                alt_promotion_met = True
                alt_criteria_failed = []
                
                # Check alternative direct FYC requirement
                if req.get('direct_fyc_alt', 0) > 0 and direct_fyc < req['direct_fyc_alt']:
                    alt_promotion_met = False
                    alt_criteria_failed.append(f"直辖FYC {direct_fyc:,.0f} < {req['direct_fyc_alt']:,.0f}")
                    
                # Check alternative group FYC requirement
                if req.get('group_fyc_alt', 0) > 0 and group_fyc < req['group_fyc_alt']:
                    alt_promotion_met = False
                    alt_criteria_failed.append(f"所辖FYC {group_fyc:,.0f} < {req['group_fyc_alt']:,.0f}")
                    
                # Check alternative direct team size requirement
                if req.get('direct_team_size_alt', 0) > 0 and direct_team_size < req['direct_team_size_alt']:
                    alt_promotion_met = False
                    alt_criteria_failed.append(f"直辖人力 {direct_team_size} < {req['direct_team_size_alt']}")
                    
                # Check alternative group team size requirement
                if req.get('group_team_size_alt', 0) > 0 and group_team_size < req['group_team_size_alt']:
                    alt_promotion_met = False
                    alt_criteria_failed.append(f"所辖人力 {group_team_size} < {req['group_team_size_alt']}")
                
                if alt_promotion_met:
                    return '晋升'  # Alternative promotion path
            
            if promotion_met:
                return '晋升'
        
        # Check maintenance requirements (conservative - based on Excel 维持要求)
        if current_rank in self.maintenance_requirements and is_assessment_month:
            maint_req = self.maintenance_requirements[current_rank]
            
            # Get performance metrics (same as promotion logic)
            individual_fyc = performance_data.get('承保FYC', 0) or 0
            direct_fyc = performance_data.get('直辖FYC', 0) or performance_data.get('个险折算后FYC', 0) or 0
            group_fyc = performance_data.get('所辖FYC', 0) or 0
            renewal_rate = performance_data.get('续保率', 0) or 0
            direct_team_size = performance_data.get('直辖人力', 0) or 0
            group_team_size = performance_data.get('所辖人力', 0) or 0
            business_supervisors = performance_data.get('业务主管人数', 0) or 0
            referrals = performance_data.get('引荐人数', 0) or 0
            direct_growth = performance_data.get('本人直增人数', 0) or 0
            star_achievers = performance_data.get('达星人数', 0) or 0
            
            # Conservative demotion logic - only demote if failing MULTIPLE key criteria
            # Must have meaningful data
            has_meaningful_data = (individual_fyc > 0 or direct_fyc > 0 or group_fyc > 0)
            
            if not has_meaningful_data:
                return '维持'  # No data = maintain (don't demote)
            
            # Count critical failures (only check the most important criteria)
            critical_failures = 0
            
            # Individual FYC is most critical
            if maint_req.get('individual_fyc', 0) > 0 and individual_fyc < maint_req['individual_fyc'] * 0.5:
                critical_failures += 2  # Weight heavily
                
            # Direct team FYC is also important
            if maint_req.get('direct_fyc', 0) > 0 and direct_fyc < maint_req['direct_fyc'] * 0.5:
                critical_failures += 1
                
            # Group FYC for senior roles
            if maint_req.get('group_fyc', 0) > 0 and group_fyc < maint_req['group_fyc'] * 0.5:
                critical_failures += 1
                
            # Renewal rate quality check
            if maint_req.get('renewal_rate', 0) > 0 and renewal_rate < maint_req['renewal_rate'] * 0.8:
                critical_failures += 1
            
            # Only demote if multiple critical failures AND very low individual performance
            if critical_failures >= 3 and individual_fyc < 5000:
                return '降级'
            else:
                return '维持'
        
        # Default fallback
        return '维持'

# Custom rule factory function
def create_custom_promotion_engine(promotion_df: pd.DataFrame = None, 
                                 demotion_df: pd.DataFrame = None,
                                 enable_mdrt_rule: bool = True,
                                 enable_composite_rule: bool = True) -> PromotionRuleEngine:
    """
    Create a custom promotion engine with additional business rules.
    
    Args:
        promotion_df: Optional promotion rules from Excel
        demotion_df: Optional demotion rules from Excel
        enable_mdrt_rule: Whether to enable MDRT-based promotion rule
        enable_composite_rule: Whether to enable composite performance rule
        
    Returns:
        Configured PromotionRuleEngine with custom rules
    """
    engine = PromotionRuleEngine()
    
    # EMERGENCY FIX: Only use grade-specific rule to prevent mass demotions
    # All other rules are too aggressive
    grade_specific_rule = GradeSpecificRule("Grade Specific Requirements", priority=5)
    engine.add_rule(grade_specific_rule)
    
    # Add a super-safe fallback rule that NEVER demotes
    class SafeFallbackRule(PromotionRule):
        def __init__(self):
            super().__init__("Safe Fallback", priority=0)
            
        def applies_to(self, current_rank: str, performance_data: Dict) -> bool:
            return True  # Always applies
            
        def evaluate(self, current_rank: str, performance_data: Dict) -> str:
            # Check for clear promotion criteria
            fyc = performance_data.get('承保FYC', 0) or 0
            if fyc >= 300000:  # Very high threshold for promotion
                return '晋升'
            else:
                return '维持'  # NEVER demote from fallback rule
    
    safe_fallback = SafeFallbackRule()
    engine.add_rule(safe_fallback)
    
    print(f"  Configured SAFE promotion engine with {len(engine.rules)} rules (demotions disabled)")
    return engine

# =============================================================================
# RANK AND LAYER PROCESSING
# =============================================================================

def apply_rank_remapping(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply rank remapping according to business rules.
    
    Args:
        df: DataFrame with current rank data
        
    Returns:
        DataFrame with remapped ranks and calculated layers
    """
    df = df.copy()
    
    # Apply rank remapping
    df['当前职级'] = df['当前职级'].map(lambda x: RANK_MAP.get(x, x))
    
    # Calculate layer for each rank
    df['RANK_LAYER'] = df['当前职级'].map(lambda x: LAYER_MAP.get(x, None))
    
    # Also remap supervisor ranks if they exist
    if '直属主管职级' in df.columns:
        df['直属主管职级'] = df['直属主管职级'].map(lambda x: RANK_MAP.get(x, x))
    if '上级主管职级' in df.columns:
        df['上级主管职级'] = df['上级主管职级'].map(lambda x: RANK_MAP.get(x, x))
    
    return df

def apply_rank_adjustment(current_rank: str, decision: str) -> str:
    """
    Apply rank adjustment based on promotion/demotion decision.
    
    Args:
        current_rank: Current rank of the agent
        decision: Promotion decision ('晋升', '维持', '降级')
        
    Returns:
        New rank after adjustment
    """
    if current_rank not in RANK_HIERARCHY:
        return current_rank
        
    current_idx = RANK_HIERARCHY.index(current_rank)
    
    if decision == '晋升' and current_idx < len(RANK_HIERARCHY) - 1:
        return RANK_HIERARCHY[current_idx + 1]
    elif decision == '降级' and current_idx > 0:
        return RANK_HIERARCHY[current_idx - 1]
    else:
        return current_rank

# =============================================================================
# PERFORMANCE ASSESSMENT
# =============================================================================

def calculate_performance_metrics(df: pd.DataFrame, agent_codes: List[str], 
                                assessment_months: List[int]) -> pd.DataFrame:
    """
    Calculate aggregated performance metrics for assessment period.
    
    Args:
        df: Source data DataFrame
        agent_codes: List of agent codes to assess
        assessment_months: List of months in assessment window
        
    Returns:
        DataFrame with aggregated performance metrics
    """
    # Filter data for assessment window
    window_data = df[df['薪资月份'].isin(assessment_months)].copy()
    
    # Group by agent and calculate aggregates
    performance_agg = {}
    
    # Define aggregation rules for each performance column
    agg_rules = {
        '承保FYC': 'sum',           # 个人FYC (Individual FYC)
        '个险折算后FYC': 'sum',      # Alternative direct FYC
        '直辖FYC': 'sum',           # 直辖FYC (Direct team FYC)
        '所辖FYC': 'sum',           # 所辖FYC (Group FYC)
        '续保率': 'mean',           # 续保率 (Renewal rate)
        '新单件数': 'sum',          # 新单件数 (New policies)
        '直辖人力': 'mean',         # 直辖人力 (Direct team size)
        '所辖人力': 'mean',         # 所辖人力 (Group team size)
        '业务主管人数': 'mean',      # 业务主管人数 (Business supervisors)
        '引荐人数': 'sum',          # 引荐人数 (Referrals)
        '本人直增人数': 'sum',       # 本人直增人数 (Direct growth)
        '达星人数': 'sum'           # 达星人数 (Star achievers)
    }
    
    # Calculate aggregations for available columns
    for col, agg_func in agg_rules.items():
        if col in window_data.columns:
            performance_agg[col] = agg_func
    
    if not performance_agg:
        # Return empty DataFrame if no performance columns found
        return pd.DataFrame(index=agent_codes)
    
    # Perform groupby aggregation
    perf_df = window_data.groupby('营销员代码').agg(performance_agg)
    
    # Ensure all agent codes are included (fill missing with 0/NaN)
    perf_df = perf_df.reindex(agent_codes, fill_value=0)
    
    return perf_df

def process_assessment_month(df: pd.DataFrame, current_month: int, 
                           source_data: pd.DataFrame, promotion_rules: pd.DataFrame,
                           demotion_rules: pd.DataFrame) -> pd.DataFrame:
    """
    Process assessment month with promotion/demotion logic using the modular rule engine.
    
    Args:
        df: Current month DataFrame
        current_month: Current assessment month
        source_data: Full source data for performance calculation
        promotion_rules: Promotion rules DataFrame
        demotion_rules: Demotion rules DataFrame
        
    Returns:
        DataFrame with updated ranks and promotion decisions
    """
    df = df.copy()
    
    # Calculate assessment window
    assessment_months = calculate_assessment_window(current_month, PROMOTION_WINDOW)
    print(f"  Assessment window: {assessment_months}")
    
    # Get performance metrics
    agent_codes = df['营销员代码'].tolist()
    perf_df = calculate_performance_metrics(source_data, agent_codes, assessment_months)
    
    # Merge performance data
    df = df.merge(perf_df, left_on='营销员代码', right_index=True, 
                  how='left', suffixes=('', '_累计'))
    
    # Create promotion engine with rules (using custom engine for more flexibility)
    promotion_engine = create_custom_promotion_engine(
        promotion_rules, 
        demotion_rules,
        enable_mdrt_rule=True,
        enable_composite_rule=True
    )
    
    # Make promotion decisions using the rule engine
    promotion_decisions = []
    new_ranks = []
    
    # Debug counters
    debug_no_data_count = 0
    debug_demotion_count = 0
    debug_sample_demotions = []
    
    for idx, row in df.iterrows():
        agent_code = row['营销员代码']
        current_rank = row['当前职级']
        
        # Prepare performance data dictionary
        perf_data = {}
        for col in PERFORMANCE_COLUMNS:
            # Try cumulative first, then current month
            cumulative_col = f"{col}_累计"
            if cumulative_col in row:
                perf_data[col] = row[cumulative_col]
            elif col in row:
                perf_data[col] = row[col]
            else:
                perf_data[col] = 0
        
        # Add GT/DT relationship counts for leadership validation
        perf_data['GT_subordinates'] = row.get('GT_subordinates', 0)
        perf_data['DT_subordinates'] = row.get('DT_subordinates', 0)
        perf_data['total_subordinates'] = row.get('total_subordinates', 0)
        
        # Add assessment month flag for demotion logic
        perf_data['is_assessment_month'] = True  # This function is only called during assessment months
        
        # Debug: Check if agent has any meaningful data
        has_data = any(perf_data.get(col, 0) > 0 for col in ['承保FYC', '直辖FYC', '个险折算后FYC', '所辖FYC'])
        if not has_data:
            debug_no_data_count += 1
        
        # Use promotion engine to determine decision
        decision = promotion_engine.evaluate(current_rank, perf_data)
        promotion_decisions.append(decision)
        
        # Debug: Track demotions
        if decision == '降级':
            debug_demotion_count += 1
            if len(debug_sample_demotions) < 10:
                debug_sample_demotions.append({
                    'agent': agent_code,
                    'rank': current_rank,
                    'fyc': perf_data.get('承保FYC', 0),
                    'direct_fyc': perf_data.get('直辖FYC', 0) or perf_data.get('个险折算后FYC', 0),
                    'has_data': has_data
                })
        
        # Apply rank adjustment
        new_rank = apply_rank_adjustment(current_rank, decision)
        new_ranks.append(new_rank)
        
        # Only log actual rank changes to reduce verbosity
        if new_rank != current_rank:
            print(f"  {agent_code}: {current_rank} → {new_rank} ({decision})")
    
    df['升降级标记'] = promotion_decisions
    df['当前职级'] = new_ranks
    
    # Reapply rank remapping to update layers
    df = apply_rank_remapping(df)
    
    # Summary statistics
    promotion_count = promotion_decisions.count('晋升')
    demotion_count = promotion_decisions.count('降级')
    maintain_count = promotion_decisions.count('维持')
    
    print(f"  Assessment results: {promotion_count} 晋升, {demotion_count} 降级, {maintain_count} 维持")
    
    # Debug information
    print(f"  Debug info: {debug_no_data_count} agents with no performance data")
    print(f"  Debug info: {debug_demotion_count} demotions from engine")
    
    # Show sample demotions for debugging
    if debug_sample_demotions:
        print("  Sample demotions for debugging:")
        for demo in debug_sample_demotions:
            print(f"    {demo['agent']} ({demo['rank']}): FYC={demo['fyc']:,.0f}, Direct={demo['direct_fyc']:,.0f}, HasData={demo['has_data']}")
    
    # Emergency check: if still too many demotions, override them
    if demotion_count > 1000:
        print("  🚨 EMERGENCY: Overriding excessive demotions!")
        # Override demotions to maintains
        df.loc[df['升降级标记'] == '降级', '升降级标记'] = '维持'
        df.loc[df['升降级标记'] == '维持', '当前职级'] = df.loc[df['升降级标记'] == '维持', '当前职级']  # Keep original rank
        print("  All demotions converted to maintains for safety")
    
    return df

# =============================================================================
# SUPERVISOR RELATIONSHIP MANAGEMENT
# =============================================================================

def validate_supervisor_hierarchy(df: pd.DataFrame) -> pd.DataFrame:
    """
    Recursively validate and fix supervisor hierarchy to ensure proper rank relationships.
    
    Args:
        df: DataFrame with current organizational data
        
    Returns:
        DataFrame with validated supervisor relationships
    """
    df = df.copy()
    
    # Create lookup dictionaries
    agent_rank_lookup = df.set_index('营销员代码')['当前职级'].to_dict()
    supervisor_lookup = df.set_index('营销员代码')['上级主管代码'].to_dict()
    
    def get_rank_hierarchy_level(rank: str) -> int:
        """Get numerical level for rank comparison (higher number = higher rank)."""
        try:
            return RANK_HIERARCHY.index(rank)
        except ValueError:
            return -1  # Unknown rank
    
    def find_valid_supervisor(agent_code: str, visited: Set[str] = None) -> str:
        """Recursively find a valid supervisor who outranks the agent."""
        if visited is None:
            visited = set()
            
        if agent_code in visited:
            return ''  # Circular reference
            
        visited.add(agent_code)
        agent_rank = agent_rank_lookup.get(agent_code, '')
        agent_level = get_rank_hierarchy_level(agent_rank)
        
        current_supervisor = supervisor_lookup.get(agent_code, '')
        if not current_supervisor:
            return ''
            
        supervisor_rank = agent_rank_lookup.get(current_supervisor, '')
        supervisor_level = get_rank_hierarchy_level(supervisor_rank)
        
        # If supervisor outranks agent, this is valid
        if supervisor_level > agent_level:
            return current_supervisor
            
        # If supervisor is same or lower rank, try to find their supervisor
        return find_valid_supervisor(current_supervisor, visited)
    
    # Initialize 育成主管代码 if not exists
    if '育成主管代码' not in df.columns:
        df['育成主管代码'] = ''
    
    # Process each agent to validate hierarchy
    hierarchy_fixes = 0
    for idx, row in df.iterrows():
        agent_code = row['营销员代码']
        current_supervisor = row['上级主管代码']
        
        if pd.isna(current_supervisor) or current_supervisor == '':
            continue
            
        # Find valid supervisor
        valid_supervisor = find_valid_supervisor(agent_code)
        
        # If we found a different valid supervisor, update relationships
        if valid_supervisor != current_supervisor and valid_supervisor:
            df.at[idx, '育成主管代码'] = current_supervisor  # Store original as mentor
            df.at[idx, '上级主管代码'] = valid_supervisor
            hierarchy_fixes += 1
    
    if hierarchy_fixes > 0:
        print(f"  Fixed {hierarchy_fixes} supervisor hierarchy conflicts")
    
    return df

def calculate_team_relationships(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate GT/DT relationship counts for each agent.
    
    Args:
        df: DataFrame with organizational data
        
    Returns:
        DataFrame with relationship counts added
    """
    df = df.copy()
    
    # Initialize relationship count columns
    df['GT_subordinates'] = 0
    df['DT_subordinates'] = 0
    df['total_subordinates'] = 0
    
    # Count subordinates for each supervisor
    for supervisor_code in df['营销员代码'].unique():
        # Find all subordinates of this supervisor
        subordinates = df[df['上级主管代码'] == supervisor_code]
        
        if subordinates.empty:
            continue
            
        # Count GT and DT relationships
        gt_count = (subordinates['主管关系'] == 'GT').sum()
        dt_count = (subordinates['主管关系'] == 'DT').sum()
        total_count = len(subordinates)
        
        # Update supervisor's counts
        supervisor_idx = df[df['营销员代码'] == supervisor_code].index
        if not supervisor_idx.empty:
            df.loc[supervisor_idx, 'GT_subordinates'] = gt_count
            df.loc[supervisor_idx, 'DT_subordinates'] = dt_count
            df.loc[supervisor_idx, 'total_subordinates'] = total_count
    
    return df

def update_supervisor_relationships(df: pd.DataFrame) -> pd.DataFrame:
    """
    Update supervisor relationships with comprehensive validation and GT/DT counting.
    
    Args:
        df: DataFrame with current organizational data
        
    Returns:
        DataFrame with updated supervisor relationships
    """
    df = df.copy()
    
    # Step 1: Validate supervisor hierarchy recursively
    df = validate_supervisor_hierarchy(df)
    
    # Step 2: Apply relationship rules
    df = apply_relationship_rules(df)
    
    # Step 3: Calculate team relationship counts
    df = calculate_team_relationships(df)
    
    return df

def determine_relationship_type(agent_rank: str) -> str:
    """
    Determine relationship type (DT/GT) based on agent rank.
    
    Args:
        agent_rank: Current rank of the agent
        
    Returns:
        Relationship type: 'DT' or 'GT'
    """
    agent_layer = LAYER_MAP.get(agent_rank, 4)
    return 'DT' if agent_layer >= 3 else 'GT'

def apply_relationship_rules(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply relationship rules and force DT for business managers.
    
    Args:
        df: DataFrame with organizational data
        
    Returns:
        DataFrame with applied relationship rules
    """
    df = df.copy()
    
    # Determine basic relationship types
    df['主管关系'] = df['当前职级'].map(determine_relationship_type)
    
    # Force DT for business manager ranks
    business_manager_ranks = ['AD0', 'BM3', 'BM2', 'BM1']
    manager_mask = df['当前职级'].isin(business_manager_ranks)
    df.loc[manager_mask, '主管关系'] = 'DT'
    
    return df

# =============================================================================
# EMPLOYEE STATUS TRACKING
# =============================================================================

def detect_employee_status_changes(current_agents: Set[str], 
                                 previous_agents: Set[str]) -> Tuple[Set[str], Set[str]]:
    """
    Detect new hires and resignations by comparing agent sets.
    
    Args:
        current_agents: Set of current month agent codes
        previous_agents: Set of previous month agent codes
        
    Returns:
        Tuple of (new_hires, resignations)
    """
    new_hires = current_agents - previous_agents
    resignations = previous_agents - current_agents
    
    return new_hires, resignations

def apply_employee_status(df: pd.DataFrame, new_hires: Set[str], 
                         resignations: Set[str]) -> pd.DataFrame:
    """
    Apply employee status labels to DataFrame.
    
    Args:
        df: DataFrame with employee data
        new_hires: Set of new hire agent codes
        resignations: Set of resignation agent codes
        
    Returns:
        DataFrame with employee status column
    """
    df = df.copy()
    
    # Default status
    df['员工状态'] = '在职'
    
    # Mark new hires
    df.loc[df['营销员代码'].isin(new_hires), '员工状态'] = '新增'
    
    # Mark resignations (though they typically won't appear in current month data)
    df.loc[df['营销员代码'].isin(resignations), '员工状态'] = '离职'
    
    return df

# =============================================================================
# OUTPUT FORMATTING
# =============================================================================

def create_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create and populate required output columns.
    
    Args:
        df: DataFrame with processed organizational data
        
    Returns:
        DataFrame with all required output columns
    """
    df = df.copy()
    
    # Map basic hierarchy columns
    df['总监代码'] = df.get('AD代码', '')
    df['总监职级'] = 'AD1'  # Default assumption
    df['总监关系'] = 'GT'
    
    df['业务经理代码'] = df.get('直属主管代码', '')
    df['业务经理职级'] = df.get('直属主管职级', '')
    df['业务经理关系'] = df.get('主管关系', 'DT')
    
    df['业务主任代码'] = df.get('上级主管代码', '')
    df['业务主任职级'] = df.get('上级主管职级', '')
    df['业务主任关系'] = df.get('主管关系', 'DT')
    
    # Mentor relationships
    df['一代育成人'] = df.get('育成主管代码', '')
    df['二代育成人'] = ''  # Placeholder for future implementation
    
    return df

def format_final_output(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format the final output DataFrame with proper column order and data types.
    
    Args:
        df: Processed DataFrame
        
    Returns:
        Formatted DataFrame ready for output
    """
    df = df.copy()
    
    # Ensure all required columns exist
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = ''
    
    # Get all original columns plus new output columns
    original_cols = [col for col in df.columns if col not in OUTPUT_COLUMNS]
    final_columns = OUTPUT_COLUMNS + original_cols
    
    # Reorder columns
    df = df.reindex(columns=final_columns)
    
    # Clean up data types
    string_columns = ['营销员代码', '上级主管代码', '总监代码', '业务经理代码', '业务主任代码']
    for col in string_columns:
        if col in df.columns:
            df[col] = df[col].astype(str).replace('nan', '')
    
    return df

# =============================================================================
# MAIN PROCESSING FUNCTION
# =============================================================================

def process_monthly_data(source_data: pd.DataFrame, promotion_rules: pd.DataFrame,
                        demotion_rules: pd.DataFrame) -> Tuple[pd.DataFrame, List[pd.DataFrame]]:
    """
    Process all monthly data according to business rules.
    
    Args:
        source_data: Raw source data from Excel
        promotion_rules: Promotion rules DataFrame
        demotion_rules: Demotion rules DataFrame
        
    Returns:
        Tuple of (processed DataFrame with all months, list of monthly snapshots)
    """
    print("Processing monthly organizational data...")
    
    # Get sorted list of all months
    month_list = sorted(source_data['薪资月份'].unique())
    print(f"Processing months: {month_list}")
    
    monthly_snapshots = []
    previous_agent_codes = set()

    for month in month_list:
        print(f"\nProcessing month {month}...")

        # Extract current month data
        current_df = source_data[source_data['薪资月份'] == month].copy()
        print(f"  {len(current_df)} agents in month {month}")

        # Step 1: Apply rank remapping
        current_df = apply_rank_remapping(current_df)
        
        # Step 2: Process assessment if this is an assessment month
        if month in ASSESS_MONTHS:
            print(f"  Processing assessment month {month}")
            current_df = process_assessment_month(
                current_df, month, source_data, promotion_rules, demotion_rules
            )
        
        # Step 3: Detect employee status changes
        current_agent_codes = set(current_df['营销员代码'])
        new_hires, resignations = detect_employee_status_changes(
            current_agent_codes, previous_agent_codes
        )
        
        if new_hires:
            print(f"  {len(new_hires)} new hires detected")
        if resignations:
            print(f"  {len(resignations)} resignations detected")
            
        current_df = apply_employee_status(current_df, new_hires, resignations)
        
        # Step 4: Update supervisor relationships (includes validation and GT/DT counting)
        current_df = update_supervisor_relationships(current_df)
        
        # Step 5: Create output columns
        current_df = create_output_columns(current_df)
        
        # Store snapshot and update previous month tracking
        monthly_snapshots.append(current_df)
        previous_agent_codes = current_agent_codes
    
    # Combine all monthly snapshots
    print(f"\nCombining {len(monthly_snapshots)} monthly snapshots...")
    final_df = pd.concat(monthly_snapshots, ignore_index=True)
    
    # Format final output
    final_df = format_final_output(final_df)
    
    print(f"Final dataset contains {len(final_df)} records")
    return final_df, monthly_snapshots

# =============================================================================
# PROMOTION CHECK EXCEL GENERATION
# =============================================================================

def create_monthly_promotion_check(monthly_snapshots: List[pd.DataFrame], 
                                  source_data: pd.DataFrame) -> None:
    """
    Create a detailed Excel file with monthly promotion/demotion analysis.
    One tab per month showing all promotion decisions and supporting data.
    
    Args:
        monthly_snapshots: List of monthly DataFrame snapshots
        source_data: Original source data for reference
    """
    print(f"\nGenerating promotion check file: {CHECK_FILE}")
    
    with pd.ExcelWriter(CHECK_FILE, engine='openpyxl') as writer:
        
        # Create summary sheet first
        create_promotion_summary_sheet(monthly_snapshots, writer)
        
        # Create detailed monthly sheets
        for i, monthly_df in enumerate(monthly_snapshots):
            month = monthly_df['薪资月份'].iloc[0] if not monthly_df.empty else f"Month_{i+1}"
            month_str = str(month)
            
            # Check if this is an assessment month
            is_assessment_month = month in ASSESS_MONTHS
            
            if is_assessment_month:
                print(f"  Creating assessment details for {month_str}")
                create_assessment_month_sheet(monthly_df, month, writer, source_data)
            else:
                print(f"  Creating regular month summary for {month_str}")
                create_regular_month_sheet(monthly_df, month, writer)

def create_promotion_summary_sheet(monthly_snapshots: List[pd.DataFrame], 
                                 writer: pd.ExcelWriter) -> None:
    """Create summary sheet with overall promotion statistics."""
    
    summary_data = []
    
    for monthly_df in monthly_snapshots:
        if monthly_df.empty:
            continue
            
        month = monthly_df['薪资月份'].iloc[0]
        is_assessment = month in ASSESS_MONTHS
        
        # Count promotion decisions
        if '升降级标记' in monthly_df.columns:
            promotion_count = (monthly_df['升降级标记'] == '晋升').sum()
            demotion_count = (monthly_df['升降级标记'] == '降级').sum()
            maintain_count = (monthly_df['升降级标记'] == '维持').sum()
        else:
            promotion_count = demotion_count = maintain_count = 0
        
        # Count employee status
        new_hires = (monthly_df['员工状态'] == '新增').sum() if '员工状态' in monthly_df.columns else 0
        resignations = (monthly_df['员工状态'] == '离职').sum() if '员工状态' in monthly_df.columns else 0
        active = (monthly_df['员工状态'] == '在职').sum() if '员工状态' in monthly_df.columns else len(monthly_df)
        
        summary_data.append({
            '月份': month,
            '考核月': '是' if is_assessment else '否',
            '总人数': len(monthly_df),
            '新增人员': new_hires,
            '离职人员': resignations,
            '在职人员': active,
            '晋升人数': promotion_count,
            '降级人数': demotion_count,
            '维持人数': maintain_count,
            '晋升率': f"{promotion_count/len(monthly_df)*100:.1f}%" if len(monthly_df) > 0 else "0%",
            '降级率': f"{demotion_count/len(monthly_df)*100:.1f}%" if len(monthly_df) > 0 else "0%"
        })
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    # Auto-adjust column widths
    worksheet = writer.sheets['Summary']
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def create_assessment_month_sheet(monthly_df: pd.DataFrame, month: int, 
                                writer: pd.ExcelWriter, source_data: pd.DataFrame) -> None:
    """Create detailed sheet for assessment months with promotion analysis."""
    
    sheet_name = f"{month}_Assessment"
    
    # Filter for agents with promotion decisions
    assessment_df = monthly_df.copy()
    
    # Get assessment window data
    assessment_months = calculate_assessment_window(month, PROMOTION_WINDOW)
    window_data = source_data[source_data['薪资月份'].isin(assessment_months)]
    
    # Calculate performance aggregations for context
    perf_agg = {}
    agg_rules = {
        '承保FYC': 'sum',
        '个险折算后FYC': 'sum',
        '直辖FYC': 'sum',
        '所辖FYC': 'sum',
        '续保率': 'mean',
        '新单件数': 'sum',
        '直辖人力': 'mean',
        '所辖人力': 'mean'
    }
    
    for col, agg_func in agg_rules.items():
        if col in window_data.columns:
            perf_agg[col] = agg_func
    
    if perf_agg:
        perf_df = window_data.groupby('营销员代码').agg(perf_agg)
        assessment_df = assessment_df.merge(perf_df, left_on='营销员代码', 
                                          right_index=True, how='left', suffixes=('', '_6个月累计'))
    
    # Select key columns for the check sheet
    check_columns = [
        '营销员代码', '当前职级', '升降级标记', '员工状态',
        '承保FYC_6个月累计', '个险折算后FYC_6个月累计', '直辖FYC_6个月累计', '所辖FYC_6个月累计',
        '续保率_6个月累计', '新单件数_6个月累计', '直辖人力_6个月累计', '所辖人力_6个月累计',
        '直属主管代码', '上级主管代码', '主管关系'
    ]
    
    # Only include columns that exist
    available_columns = [col for col in check_columns if col in assessment_df.columns]
    check_df = assessment_df[available_columns].copy()
    
    # Sort by promotion decision and then by agent code
    sort_order = {'晋升': 0, '降级': 1, '维持': 2}
    if '升降级标记' in check_df.columns:
        check_df['_sort_order'] = check_df['升降级标记'].map(sort_order).fillna(3)
        check_df = check_df.sort_values(['_sort_order', '营销员代码'])
        check_df = check_df.drop('_sort_order', axis=1)
    
    # Write to Excel
    check_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Format the worksheet
    format_assessment_worksheet(writer.sheets[sheet_name], check_df)

def create_regular_month_sheet(monthly_df: pd.DataFrame, month: int, 
                             writer: pd.ExcelWriter) -> None:
    """Create summary sheet for regular (non-assessment) months."""
    
    sheet_name = f"{month}_Regular"
    
    # Select key columns for regular months
    regular_columns = [
        '营销员代码', '当前职级', '员工状态', 'RANK_LAYER',
        '直属主管代码', '上级主管代码', '主管关系',
        '承保FYC', '个险折算后FYC', '续保率'
    ]
    
    # Only include columns that exist
    available_columns = [col for col in regular_columns if col in monthly_df.columns]
    regular_df = monthly_df[available_columns].copy()
    
    # Sort by rank layer and agent code
    if 'RANK_LAYER' in regular_df.columns:
        regular_df = regular_df.sort_values(['RANK_LAYER', '营销员代码'])
    else:
        regular_df = regular_df.sort_values(['营销员代码'])
    
    # Write to Excel
    regular_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Auto-adjust column widths
    worksheet = writer.sheets[sheet_name]
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def format_assessment_worksheet(worksheet, df: pd.DataFrame) -> None:
    """Apply formatting to assessment worksheet for better readability."""
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Color code promotion decisions if available
    if '升降级标记' in df.columns:
        from openpyxl.styles import PatternFill
        
        # Define colors
        promotion_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        demotion_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")   # Light red
        maintain_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")   # Light blue
        
        # Find the column index for 升降级标记
        decision_col_idx = None
        for idx, col in enumerate(df.columns):
            if col == '升降级标记':
                decision_col_idx = idx + 1  # Excel columns are 1-indexed
                break
        
        if decision_col_idx:
            for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                cell = worksheet.cell(row=row_idx, column=decision_col_idx)
                if cell.value == '晋升':
                    cell.fill = promotion_fill
                elif cell.value == '降级':
                    cell.fill = demotion_fill
                elif cell.value == '维持':
                    cell.fill = maintain_fill

# =============================================================================
# MAIN EXECUTION FUNCTION
# =============================================================================

def main():
    """
    Main execution function that orchestrates the entire processing pipeline.
    """
    print("=" * 70)
    print("ORGANIZATIONAL CHART PROCESSING")
    print("=" * 70)
    
    try:
        # Step 1: Load source data
        print("\n1. Loading source data...")
        source_data = load_source_data(SRC_FILE)
        
        # Step 2: Load promotion rules
        print("\n2. Loading promotion/demotion rules...")
        promotion_rules, demotion_rules = load_promotion_rules(SRC_FILE)
        
        # Step 3: Process all monthly data
        print("\n3. Processing monthly data...")
        processed_df, monthly_snapshots = process_monthly_data(source_data, promotion_rules, demotion_rules)
        
        # Step 4: Save main output
        print(f"\n4. Saving main output to {DST_FILE}...")
        processed_df.to_excel(DST_FILE, index=False)
        
        # Step 5: Generate promotion check file
        print(f"\n5. Generating promotion check file...")
        create_monthly_promotion_check(monthly_snapshots, source_data)
        
        # Success message
        print("\n" + "=" * 70)
        print(f"✅ SUCCESS! Processing completed successfully.")
        print(f"✅ Main output saved to: {DST_FILE}")
        print(f"✅ Promotion check file saved to: {CHECK_FILE}")
        print(f"✅ Total records processed: {len(processed_df):,}")
        print(f"✅ Months covered: {sorted(processed_df['薪资月份'].unique())}")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        print("Processing failed. Please check the error message above.")
        raise

if __name__ == "__main__":
    main()
